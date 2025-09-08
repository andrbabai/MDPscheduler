"""Schedule builder: download XLSX and convert to ICS."""
from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
import re
import tempfile
import uuid

import openpyxl
import requests
from icalendar import Calendar, Event
from dateutil import tz
from pydantic import BaseModel
import yaml

DAYS = [
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
]


class ScanConfig(BaseModel):
    max_header_rows: int = 8
    date_scan_up: int = 8
    max_row: int | None = None
    # New scanning strategy settings (row-wise parsing)
    header_min_row: int = 1           # first header row (inclusive)
    header_max_row: int = 5           # last header row (inclusive)
    time_row: int = 2                 # row with time headers per column
    first_event_row: int = 6          # first row with event data
    col_start: int = 2                # B
    col_end: int = 14                 # N
    default_duration_minutes: int = 90


class Config(BaseModel):
    public_link: str
    sheet_name: str | None = None
    timezone: str = "Europe/Moscow"
    year: int
    uid_prefix: str = ""
    scan: ScanConfig = ScanConfig()


TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})")
SINGLE_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})$")
DATE_RE = re.compile(r"(\d{1,2})[.,](\d{1,2})")

# Keywords indicating a special event that should be highlighted
# (matched case-insensitively within summary/description)
SPECIAL_KEYWORDS = {
    "дедлайн",
    "защита",
    "зачет",
    "зачёт",
    "экзамен",
    "начало",
}

HIGHLIGHT_COLOR = "#fadadd"


def read_config(path: str | Path) -> Config:
    """Read configuration from YAML file."""
    with open(path, "r", encoding="utf8") as f:
        data = yaml.safe_load(f)
    return Config.model_validate(data)


def download_xlsx(public_link: str, dest: Path) -> Path:
    """Download public XLSX from Yandex.Disk to dest."""
    api = "https://cloud-api.yandex.net/v1/disk/public/resources/download"
    r = requests.get(api, params={"public_key": public_link})
    r.raise_for_status()
    href = r.json()["href"]
    r2 = requests.get(href)
    r2.raise_for_status()
    dest.write_bytes(r2.content)
    return dest


def _cell_value(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            top = ws.cell(rng.min_row, rng.min_col)
            return top.value
    return None


def _is_top_left(ws, row: int, col: int) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return row == rng.min_row and col == rng.min_col
    return True


def _find_date(ws, row: int, col: int, cfg: Config) -> date:
    """Scan upward to find a date for a cell, supporting merged cells and xlsx date types.

    Looks up within ``cfg.scan.date_scan_up`` rows for:
    - a datetime/date cell (returns its date), or
    - a string containing DD.MM or DD,MM pattern.
    """
    for r in range(row, max(row - cfg.scan.date_scan_up, 0), -1):
        val = _cell_value(ws, r, col)
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            m = DATE_RE.search(val)
            if m:
                d, mth = int(m.group(1)), int(m.group(2))
                return date(cfg.year, mth, d)
    raise ValueError(f"Date not found for cell {row},{col}")


def _iter_events(ws, cfg: Config):
    # Respect max_row limit if configured
    max_row = cfg.scan.max_row if getattr(cfg, "scan", None) and cfg.scan.max_row else ws.max_row

    tzinfo = tz.gettz(cfg.timezone)

    # Helper: scan upward in a column to find time range from any header cell
    def _scan_time_up(row: int, col: int) -> tuple[time, time | None] | None:
        for r2 in range(row, 0, -1):
            v = _cell_value(ws, r2, col)
            if isinstance(v, str):
                m = TIME_RE.search(v)
                if m:
                    start = time(int(m.group(1)), int(m.group(2)))
                    end = time(int(m.group(3)), int(m.group(4)))
                    return (start, end)
                m2 = SINGLE_TIME_RE.match(v.strip())
                if m2:
                    start = time(int(m2.group(1)), int(m2.group(2)))
                    # No end time known here
                    return (start, None)  # type: ignore
        return None

    # Helper: scan upward in a column to find a date (accept xlsx date/datetime or DD.MM in text)
    def _scan_date_up(row: int, col: int, cfg: Config) -> date:
        for r2 in range(row, max(row - cfg.scan.date_scan_up, 0), -1):
            v = _cell_value(ws, r2, col)
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, str):
                m = DATE_RE.search(v)
                if m:
                    d, mth = int(m.group(1)), int(m.group(2))
                    return date(cfg.year, mth, d)
        # fallback to original behavior (scan up and error)
        return _find_date(ws, row, col, cfg)

    def _split_summary_desc(text: str) -> tuple[str, str]:
        import re as _re
        # Convert long sequences of spaces/tabs into newlines
        text = _re.sub(r"[\t ]{3,}", "\n", text)
        # Normalize lines, drop numeric-only lines from description
        parts = [ln.strip() for ln in text.splitlines()]
        clean = [ln for ln in parts if ln]
        # First non-empty line as summary
        summary = clean[0] if clean else ""
        # For description, exclude numeric-only lines
        desc_lines = [ln for ln in clean[1:] if not _re.fullmatch(r"\d+", ln)]
        # Move key markers (e.g., НАЧАЛО/ЗАЩИТА/ЗАЧЕТ/ЭКЗАМЕН) into summary together with first detail line
        KEY_MARKERS = {
            "начало",
            "защита отчета",
            "защита отчёта",
            "зачет с оценкой",
            "зачёт с оценкой",
            "экзамен",
            "зачет",
            "зачёт",
            "защита",
            "дедлайн",
        }
        if summary.strip().lower() in KEY_MARKERS and desc_lines:
            summary = f"{summary} — {desc_lines[0]}"
            desc_lines = desc_lines[1:]
        desc = "\n".join(desc_lines)
        return summary, desc

    def _is_special(summary: str, desc: str | None) -> bool:
        s = (summary or "").lower()
        d = (desc or "").lower()
        return any(k in s or k in d for k in SPECIAL_KEYWORDS)

    # Build per-column time mapping from header rows (prefer exact time_row)
    col_times: dict[int, tuple[time, time] | tuple[time, None]] = {}
    for c in range(cfg.scan.col_start, cfg.scan.col_end + 1):
        v = _cell_value(ws, cfg.scan.time_row, c)
        start_end: tuple[time, time] | tuple[time, None] | None = None
        if isinstance(v, str):
            m = TIME_RE.search(v)
            if m:
                start_end = (
                    time(int(m.group(1)), int(m.group(2))),
                    time(int(m.group(3)), int(m.group(4))),
                )
            else:
                m2 = SINGLE_TIME_RE.match(v.strip())
                if m2:
                    start_end = (
                        time(int(m2.group(1)), int(m2.group(2))),
                        None,
                    )
        # If not found on the preferred row, try within header rows window
        if not start_end:
            for r in range(cfg.scan.header_min_row, cfg.scan.header_max_row + 1):
                v2 = _cell_value(ws, r, c)
                if isinstance(v2, str):
                    m = TIME_RE.search(v2)
                    if m:
                        start_end = (
                            time(int(m.group(1)), int(m.group(2))),
                            time(int(m.group(3)), int(m.group(4))),
                        )
                        break
                    m2 = SINGLE_TIME_RE.match(v2.strip())
                    if m2:
                        start_end = (
                            time(int(m2.group(1)), int(m2.group(2))),
                            None,
                        )
                        break
        if start_end:
            col_times[c] = start_end

    # Fill missing end times using default duration only (events have only start)
    for c in range(cfg.scan.col_start, cfg.scan.col_end + 1):
        ts = col_times.get(c)
        if not ts:
            continue
        start, end = ts
        if end is None:
            # Default duration
            start_dt = datetime(2000, 1, 1, start.hour, start.minute)
            end_dt = start_dt + timedelta(minutes=cfg.scan.default_duration_minutes)
            end = time(end_dt.hour, end_dt.minute)
        col_times[c] = (start, end)

    # Helper: parse date from a single cell value
    def _parse_date_val(val) -> date | None:
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            m = DATE_RE.search(val)
            if m:
                d, mth = int(m.group(1)), int(m.group(2))
                return date(cfg.year, mth, d)
        return None

    # Maintain per-column current date, updated on special date rows (e.g., 5,7,9,...)
    current_date_for_col: dict[int, date] = {}

    def _update_date_mapping_for_row(r: int):
        updated = False
        for c in range(cfg.scan.col_start, cfg.scan.col_end + 1):
            v = _cell_value(ws, r, c)
            d = _parse_date_val(v)
            if d is not None:
                current_date_for_col[c] = d
                updated = True
        return updated

    # Prime mapping with any date rows before the first event row
    for r0 in range(cfg.scan.header_min_row, min(cfg.scan.first_event_row, max_row + 1)):
        _update_date_mapping_for_row(r0)

    # Row-wise parsing: iterate rows and columns B..N for event cells
    for r in range(cfg.scan.first_event_row, max_row + 1):
        # If this row contains date labels, refresh mapping first
        _update_date_mapping_for_row(r)
        for c in range(cfg.scan.col_start, cfg.scan.col_end + 1):
            val = _cell_value(ws, r, c)
            if not isinstance(val, str):
                continue
            if not val.strip():
                continue
            # Skip obvious header-like cells
            low = val.strip().lower()
            if low in DAYS:
                continue
            if TIME_RE.fullmatch(low) or SINGLE_TIME_RE.fullmatch(low):
                continue
            if not _is_top_left(ws, r, c):
                continue

            # Determine date: prefer per-column mapping from nearest date row (<= r)
            dt: date | None = current_date_for_col.get(c)
            if dt is None:
                # Fallback: scan upward in the same column
                dt = _scan_date_up(r, c, cfg)

            # Determine time from column header mapping
            tpair = col_times.get(c)
            if not tpair:
                # As a last resort, scan up for time anywhere in this column
                tmp = _scan_time_up(r, c)
                if not tmp or tmp[0] is None:  # type: ignore[index]
                    continue
                start, end = tmp  # type: ignore[misc]
            else:
                start, end = tpair  # type: ignore[misc]

            # If end time is missing, apply default duration
            if end is None:
                start_dt = datetime(2000, 1, 1, start.hour, start.minute)
                end_dt = start_dt + timedelta(minutes=cfg.scan.default_duration_minutes)
                end = time(end_dt.hour, end_dt.minute)

            summary, desc = _split_summary_desc(str(val))
            dtstart = datetime(dt.year, dt.month, dt.day, start.hour, start.minute, tzinfo=tzinfo)
            dtend = datetime(dt.year, dt.month, dt.day, end.hour, end.minute, tzinfo=tzinfo)
            uid = f"{cfg.uid_prefix}{dt.isoformat()}-{start.hour:02d}{start.minute:02d}-{uuid.uuid5(uuid.NAMESPACE_DNS, summary)}"
            event = Event()
            event.add("summary", summary)
            if desc:
                event.add("description", desc)
            event.add("dtstart", dtstart)
            event.add("dtend", dtend)
            event.add("uid", uid)
            if _is_special(summary, desc):
                try:
                    event.add("color", HIGHLIGHT_COLOR)
                except Exception:
                    event.add("X-COLOR", HIGHLIGHT_COLOR)
                event.add("categories", "highlight")
            yield event
    # End of row-wise parsing


def build_ics(cfg: Config) -> bytes:
    """Build calendar ICS from config."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        download_xlsx(cfg.public_link, Path(tmp.name))
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb[cfg.sheet_name] if cfg.sheet_name else wb.active
        cal = Calendar()
        cal.add("prodid", "-//schedics//")
        cal.add("version", "2.0")
        for ev in (_iter_events(ws, cfg) or []):
            cal.add_component(ev)
        return cal.to_ical()


def build_ics_and_events(cfg: Config) -> tuple[bytes, list[dict]]:
    """Build calendar ICS and return a JSON-serializable list of parsed events."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        download_xlsx(cfg.public_link, Path(tmp.name))
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb[cfg.sheet_name] if cfg.sheet_name else wb.active
        cal = Calendar()
        cal.add("prodid", "-//schedics//")
        cal.add("version", "2.0")
        events: list[dict] = []
        for ev in (_iter_events(ws, cfg) or []):
            cal.add_component(ev)
            # Extract fields for debugging/verification
            try:
                dtstart = ev.decoded("dtstart")
            except Exception:
                dtstart = None
            try:
                dtend = ev.decoded("dtend")
            except Exception:
                dtend = None
            summary = ev.get("summary")
            desc = ev.get("description")
            color_prop = ev.get("color") or ev.get("X-COLOR")
            events.append({
                "summary": str(summary) if summary is not None else None,
                "description": str(desc) if desc is not None else None,
                "dtstart": dtstart.isoformat() if hasattr(dtstart, "isoformat") else None,
                "dtend": dtend.isoformat() if hasattr(dtend, "isoformat") else None,
                "color": str(color_prop) if color_prop is not None else None,
            })
        return cal.to_ical(), events
